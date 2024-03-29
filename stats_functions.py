import logging
import os
import pandas as pd
import numpy as np
import scipy
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import scipy.stats as stats
import file_functions
from typing import Union, List, Dict, Tuple, Iterable


# STAGE BASED ANALYSIS FUNCTIONS


def runs_test(
    series: Iterable,
    units: str,
    spacing: int = 0,
) -> Dict[str, Union[float, int]]:
    """
    Does WW runs test for values above/below median of series.

    Args:
        series (iterable): a list of values for which to perform the Wald-Wolfowitz runs test for values below/above median

    Returns:
        A dictionary containing the following:
            number of runs
            number of expected runs (if random)
            expected standard deviation of number of runs (if random)
            Z: number of standard deviations difference between actual and expected number of run (standard deviation of num. of runs if random)
    """

    # convert to array if a list
    if isinstance(series, list):
        series = np.array(series)

    m = np.median(series)
    # omit values from series equal to the median
    test_series = [x for x in series if x != m]
    run_lengths = []
    count = 0
    num_in_sequence = 0
    for i, vals in enumerate(zip(test_series, test_series[1:])):
        x1, x2 = vals
        count += 1
        # if transition between value above median to value equal or below median, end the run
        if (x1 > m and x2 < m) or (x1 < m and x2 > m):
            run_lengths.append(count)
            count = 0
        else:
            num_in_sequence += 1
        # if on the last value, but no transition, then last value is part of current run
        if i == len(test_series) - 2:
            count += 1
            run_lengths.append(count)
            count = 0

    # total number of values (excluding median values)
    n = len(test_series)
    # num of values above median
    n_plus = sum([1 for x in test_series if x > m])
    # num of values below median
    n_minus = n - n_plus
    # expected number of runs if random
    exp_runs = ((2 * n_plus * n_minus * 1.0) / n) + 1
    # actual number of runs
    # Based of the Enginering Statistics Handbook. Removing 'runs' of one seems like it could make sense.
    num_runs = len(run_lengths)
    # standard deviation of expected num of runs if random
    exp_run_std = np.sqrt((exp_runs - 1) * (exp_runs - 2) * 1.0 / (n - 1))
    # number of standard deviations (of epxected run length) that the actual run count differs from WW expected run count
    z_diff_expected = (num_runs - exp_runs) * 1.0 / exp_run_std
    # Median length of a run
    median_run_length = np.mean(np.array(run_lengths))
    # Significance value of the absolute value of the Z statistic
    p_value = scipy.stats.norm.cdf(abs(z_diff_expected))

    if p_value >= 0.99:
        p_value = str(round(p_value, 5)) + '**'
    elif p_value >= 0.95:
        p_value = str(round(p_value, 5)) + '*'

    # get % of XSs in run > mean_run_length

    if spacing != 0:
        mean_run_length = median_run_length * spacing
        percent_xs_greater = (num_in_sequence / n) * 100
        data = {
            'Runs': num_runs,
            'Expected Runs': round(exp_runs, 2),
            'Expected Run StDev': round(exp_run_std, 2),
            'abs(Z)': abs(round(z_diff_expected, 2)),
            'p value': p_value,
            f'Percent of XS in run > {spacing}{units}': percent_xs_greater,
            f'Mean run length ({units})': round(mean_run_length, 2),
            f'Median run length ({units})': round(median_run_length * spacing, 2),
        }
    else:
        data = {
            'Runs': num_runs,
            'Expected Runs': round(exp_runs, 2),
            'Expected Run StDev': round(exp_run_std, 2),
            'abs(Z)': abs(round(z_diff_expected, 2)),
            'p value': p_value,
            f'Percent of XS in run > {spacing}{units}': percent_xs_greater,
            f'Median run length ({units})': round(median_run_length, 2),
        }
    num_runs = 0

    return data


def runs_test_to_xlsx(
    ws: Worksheet,
    gcs_df: pd.DataFrame,
    units: str,
    ws_start_coords: Tuple[int, int] = (16, 1),
    fields: List[str] = ['Ws', 'Zs', 'Ws_Zs'],
) -> openpyxl.Workbook:
    """Writes the output of the WW Runs test to a .xlsx file"""

    # get the starting row, column coordinates on the sheet
    base_row = ws_start_coords[0]
    base_col = ws_start_coords[1]

    ws.cell(row=base_row, column=base_col).value = 'Wald-Wolfowitz runs test'
    ws.cell(row=base_row + 1, column=base_col).value = 'Field:'

    gcs_df.sort_values(by=['dist_down'], inplace=True)
    spacing = int(gcs_df.iloc[1]['dist_down'] - gcs_df.iloc[0]['dist_down'])

    for count, field in enumerate(fields):
        col = base_col + 1 + count
        ws.cell(row=base_row + 1, column=base_col + 1 + count).value = field
        series = gcs_df.loc[:, [field]].squeeze()
        out_dict = runs_test(
            series,
            units=units,
            spacing=int(spacing),
        )

        # add runs test outputs
        if count == 0:
            for i, key in enumerate(out_dict.keys()):
                ws.cell(row=base_row + 2 + i, column=base_col).value = str(key)

        for j, key in enumerate(out_dict.values()):
            ws.cell(row=base_row + 2 + j, column=col).value = str(key)

    return ws


def descriptive_stats_xlxs(
    zs: Union[str, List[Union[float, int]]],
    analysis_dir: str,
    detrended_dem: str,
) -> str:
    """Runs stage based descriptive stats analysis and writes to an .xlsx file"""

    if detrended_dem == '':
        raise ValueError(
            'param:detrended_dem must be valid to find data directory locations + units!'
        )

    zs = file_functions.prep_key_zs(zs)

    # set up directories
    dem_dir = os.path.dirname(detrended_dem)
    gcs_dir = dem_dir + '\\gcs_tables'
    out_dir = analysis_dir + '\\stage_analysis'
    stats_xl = out_dir + '\\stage_descriptive_statistics.xlsx'

    if not os.path.exists(out_dir):
        os.makedirs(out_dir)

    # get units for labeling
    u = file_functions.get_label_units(detrended_dem)[0]

    # prep input data
    z_labels = [file_functions.float_keyz_format(z) + u for z in zs]

    # initiate excel workbook
    wb = Workbook()
    wb.save(stats_xl)

    list_of_fields = ['W', 'Z', 'Ws_Zs', 'Ws', 'Zs']

    for label in z_labels:
        stage_csv = gcs_dir + '\\%s_gcs_table.csv' % label
        stage_df = pd.read_csv(stage_csv)

        # create lists to store values
        means = []
        stds = []
        high = []
        low = []
        medians = []

        # add the mean and std_deviation of each field in list_of_fields to a list
        for field in list_of_fields[:3]:
            means.append(np.mean(stage_df.loc[:, field].to_numpy()))
            stds.append(np.std(stage_df.loc[:, field].to_numpy()))
            high.append(np.max(stage_df.loc[:, field].to_numpy()))
            low.append(np.min(stage_df.loc[:, field].to_numpy()))
            medians.append(np.median(stage_df.loc[:, field].to_numpy()))

        # make a new workbook sheet for each flow stage
        wb = openpyxl.load_workbook(stats_xl)
        ws = wb.create_sheet('%s stage stats' % label)

        ws.cell(row=2, column=1).value = 'MEAN'  # Setting up titles on xl
        ws.cell(row=3, column=1).value = 'STD'
        ws.cell(row=4, column=1).value = 'MAX'
        ws.cell(row=5, column=1).value = 'MIN'
        ws.cell(row=6, column=1).value = 'MEDIAN'

        for field in list_of_fields[:3]:  # add values in each field column
            field_index = int(list_of_fields.index(field))

            ws.cell(row=1, column=(2 + field_index)).value = field
            ws.cell(row=2, column=(2 + field_index)).value = means[field_index]
            ws.cell(row=3, column=(2 + field_index)).value = stds[field_index]
            ws.cell(row=4, column=(2 + field_index)).value = high[field_index]
            ws.cell(row=5, column=(2 + field_index)).value = low[field_index]
            ws.cell(row=6, column=(2 + field_index)
                    ).value = medians[field_index]

        wb.save(stats_xl)

        # run wald's runs test and add results to the flow stage sheet
        ws = runs_test_to_xlsx(
            ws,
            stage_df,
            units=u,
            ws_start_coords=(16, 1),
            fields=['Ws', 'Zs', 'Ws_Zs'],
        )

        # calculate descriptive statistics for cross-sections classified as each landform
        landform_dict = {
            -2: 'Oversized',
            -1: 'Constricted pool',
            0: 'Normal',
            1: 'Wide riffle',
            2: 'Nozzle',
        }
        codes = landform_dict.keys()

        ws['F1'].value = '*Code: -2 for oversized, -1 for constricted pool, 0 for normal channel, 1 for wide riffle, and 2 for nozzle'
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['A'].width = 16

        # 'W', 'Z', 'Ws_Zs'
        total_rows = len(stage_df.index)
        above_half_list = [0, 0, 0]
        above_1_list = [0, 0, 0]
        below_half_list = [0, 0, 0]
        below_1_list = [0, 0, 0]
        abs_above_half_list = [0, 0, 0]
        abs_above_1_list = [0, 0, 0]
        cwz_above_zero = 0

        for index, row in stage_df.iterrows():
            std = stds[2]
            if row['Ws_Zs'] >= 0:
                cwz_above_zero += 1

            if row['Ws_Zs'] >= std:
                above_1_list[2] += 1
                above_half_list[2] += 1
            elif row['Ws_Zs'] >= (0.5 * std):
                above_half_list[2] += 1

            if row['Ws_Zs'] <= -std:
                below_1_list[2] += 1
                below_half_list[2] += 1
            elif row['Ws_Zs'] <= -(0.5 * std):
                below_half_list[2] += 1

            if abs(row['Ws_Zs']) >= std:
                abs_above_1_list[2] += 1
                abs_above_half_list[2] += 1
            elif abs(row['Ws_Zs']) >= (0.5 * std):
                abs_above_half_list[2] += 1

            # List splice: ['W_s', 'Z_s']
            for field_index, field in enumerate(list_of_fields[3:]):
                if row[field] >= 1:
                    above_1_list[field_index] += 1
                    above_half_list[field_index] += 1
                elif row[field] >= 0.5:
                    above_half_list[field_index] += 1

                if row[field] <= -1:
                    below_1_list[field_index] += 1
                    below_half_list[field_index] += 1
                elif row[field] <= -0.5:
                    below_half_list[field_index] += 1

                if abs(row[field]) >= 1:
                    abs_above_1_list[field_index] += 1
                    abs_above_half_list[field_index] += 1
                elif abs(row[field]) >= 0.5:
                    abs_above_half_list[field_index] += 1

        ws.cell(row=7, column=1).value = "% >= 0.5 STD"
        ws.cell(row=8, column=1).value = "% >= 1 STD"
        ws.cell(row=9, column=1).value = "% <= -0.5 STD"
        ws.cell(row=10, column=1).value = "% <= -1 STD"
        ws.cell(row=11, column=1).value = r"% abs(value) >= 0.5 STD"
        ws.cell(row=12, column=1).value = r"% abs(value) >= 1 STD"
        ws.cell(row=14, column=1).value = "% C(Ws,Zs) > 0"
        ws.cell(row=14, column=2).value = float(
            (cwz_above_zero / total_rows) * 100
        )

        # calculates % of W, Z, and W_s_Z_s that are greater than 0.5 and 1 of their standard deviations
        for index in range(len(above_1_list)):
            above_half_percent = float(
                (above_half_list[index] / total_rows) * 100)
            above_1_percent = float((above_1_list[index] / total_rows) * 100)
            below_half_percent = float(
                (below_half_list[index] / total_rows) * 100)
            below_1_percent = float((below_1_list[index] / total_rows) * 100)
            abs_percent_above_half = float(
                (abs_above_half_list[index] / total_rows) * 100)
            abs_percent_above_1 = float(
                (abs_above_1_list[index] / total_rows) * 100)
            ws.cell(row=7, column=(2 + index)).value = above_half_percent
            ws.cell(row=8, column=(2 + index)).value = above_1_percent
            ws.cell(row=9, column=(2 + index)).value = below_half_percent
            ws.cell(row=10, column=(2 + index)).value = below_1_percent
            ws.cell(row=11, column=(2 + index)).value = abs_percent_above_half
            ws.cell(row=12, column=(2 + index)).value = abs_percent_above_1

        # calculating same descriptive stats for each landform, each table is spaced 7 cells apart
        row_num = 2
        for code in codes:
            code_df = stage_df.loc[stage_df['code'] == code, [
                'dist_down', 'W', 'Ws', 'Z', 'Zs', 'Ws_Zs']]
            ws.cell(row=row_num, column=7).value = (
                str(landform_dict[code]))  # Preparing the table
            ws.cell(row=row_num + 1, column=7).value = 'MEAN'
            ws.cell(row=row_num + 2, column=7).value = 'STD'
            ws.cell(row=row_num + 3, column=7).value = 'MAX'
            ws.cell(row=row_num + 4, column=7).value = 'MIN'
            ws.cell(row=row_num + 5, column=7).value = 'MEDIAN'
            ws.cell(row_num + 6, column=7).value = '% Abundance:'

            if len(code_df.index) == 0:
                for field in list_of_fields:
                    field_index = int(list_of_fields.index(field))
                    ws.cell(row=row_num, column=(
                        8 + field_index)).value = str(field)
                    ws.cell(row=row_num + 1, column=(8 + field_index)).value = 0
                    ws.cell(row=row_num + 2, column=(8 + field_index)).value = 0
                    ws.cell(row=row_num + 3, column=(8 + field_index)).value = 0
                    ws.cell(row=row_num + 4, column=(8 + field_index)).value = 0
                    ws.cell(row=row_num + 5, column=(8 + field_index)).value = 0

            else:
                for field in list_of_fields:
                    field_index = int(list_of_fields.index(field))
                    ws.cell(row=row_num, column=(
                        8 + field_index)).value = str(field)
                    ws.cell(row=row_num + 1, column=(8 + field_index)).value = (
                        np.mean(code_df.loc[:, field].to_numpy()))
                    ws.cell(row=row_num + 2, column=(8 + field_index)).value = (
                        np.std(code_df.loc[:, field].to_numpy()))
                    ws.cell(row=row_num + 3, column=(8 + field_index)).value = (
                        np.max(code_df.loc[:, field].to_numpy()))
                    ws.cell(row=row_num + 4, column=(8 + field_index)).value = (
                        np.min(code_df.loc[:, field].to_numpy()))
                    ws.cell(row=row_num + 5, column=(8 + field_index)).value = (
                        np.median(code_df.loc[:, field].to_numpy()))

            # Calculates % of XS with the given landform designation
            ws.cell(
                row_num + 6, column=8).value = float(code_df.shape[0] / total_rows) * 100

            row_num += 8

        wb.save(stats_xl)
        logging.info('Descriptive statistics table saved @ %s' % stats_xl)

    # remove the extra sheet
    extra_sheet = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(extra_sheet)
    wb.save(stats_xl)

    return stats_xl


# NESTING BASED ANALYSIS FUNCTIONS

def sankey_chi_squared(
    zs: Union[str, List[Union[float, int]]],
    aligned_gcs_csv: str,
    analysis_dir: str,
    detrended_dem: str,
) -> pd.DataFrame:
    """This function calculates the chi squared significance of landform transitions.

    Chi-Squares compares observed vs expected landform transitiosn, with expected
    frequencies being proportional to landform relative abundance. 
    Low p values indicate significant transition preferences.

    Returns: A DataFrame with the results of the Chi-Squares test.
    """

    # code number and corresponding landforms
    landforms_dict = {
        'Oversized': -2,
        'Const. Pool': -1,
        'Normal': 0,
        'Wide bar': 1,
        'Nozzle': 2,
    }
    landforms = list(landforms_dict.keys())

    if detrended_dem == '':
        raise ValueError(
            'param:detrended_dem must be valid to find data directory locations + units!'
        )

    zs = file_functions.prep_key_zs(zs)

    # set up directories
    out_dir = analysis_dir + '\\nesting_analysis'

    if not os.path.exists(out_dir):
        os.makedirs(out_dir)

    # get units for labeling
    u = file_functions.get_label_units(detrended_dem)[0]

    # prep flow stage labels
    z_labels = [file_functions.float_keyz_format(z) + u for z in zs]

    if len(z_labels) == 3:
        col_labels = ['base', 'bf', 'vf']
    else:
        col_labels = z_labels

    aligned_df = pd.read_csv(aligned_gcs_csv)

    out_dict = {
        'from': [],
        'to': [],
        'to_landform': [],
        'total_expected_freq': [],
        'expected_proportion': [],
    }

    # set up index to control list splicing
    start_i = 0

    # for each step-wise stage transition, calculate chi-squared test result
    for i in range(len(zs) - 1):
        lower = z_labels[i]
        low_code = f'{lower}_code'

        higher = z_labels[i + 1]
        high_code = f'{higher}_code'

        logging.info(
            f'Chi-sSquares test for landform transitions: {lower} -> {higher}'
        )

        type_df = aligned_df.dropna(
            axis=0,
            subset=[
                low_code,
                high_code,
            ],
            how='any',
        )

        total_rows = int(type_df.shape[0])

        for num in range(-2, 3):
            out_dict['from'].append(lower)
            out_dict['to'].append(higher)
            out_dict['to_landform'].append(landforms[num + 2])

            num_df = type_df.loc[
                lambda type_df: type_df[high_code] == num
            ]

            out_dict['total_expected_freq'].append(num_df.shape[0])

            out_dict['expected_proportion'].append(
                num_df.shape[0] / total_rows
            )

        for l_form in landforms:
            if i == 0:
                out_dict['from_' + l_form + '_freq'] = []
                out_dict['from_' + l_form + '_proportion'] = []
                out_dict['p_value_from_%s' % l_form] = []

            form_df = type_df.loc[
                lambda type_df: type_df[low_code] == landforms_dict[l_form]
            ]

            form_rows_count = form_df.shape[0]

            for h_form in landforms:

                sub_df = form_df.loc[
                    lambda form_df: form_df[high_code] == landforms_dict[h_form]
                ]

                freq = sub_df.shape[0]
                out_dict['from_' + l_form + '_freq'].append(freq)
                out_dict['from_' + l_form +
                         '_proportion'].append(freq / form_rows_count)

            # get the # of observed tranitions to the higher landform from the lower
            obs_freqs = np.array(
                out_dict['from_' + l_form + '_freq'][start_i: start_i + 5]
            )
            obs_props = np.array(
                out_dict['from_' + l_form +
                         '_proportion'][start_i: start_i + 5]
            )
            obs = (obs_props * sum(obs_freqs)).astype('int')

            # compare to the number of transitions one would expect based on
            expect_props = np.array(
                out_dict['expected_proportion'][start_i: start_i + 5]
            )
            expect = (expect_props * sum(obs_freqs)).astype('int')

            # make sure obs and expect are the same length
            remainder = sum(obs) - sum(expect)
            piece = int(remainder / abs(remainder))

            # allocate remainder based on largest rounding error
            # NOTE: a tiny bit of error can be introduced here!
            expect_props_remade = expect / sum(expect)
            gaps = expect_props - expect_props_remade

            if piece == 1:
                order_to_adjust = np.argsort(gaps, kind='mergesort')[::-1]
            elif piece == -1:
                order_to_adjust = np.argsort(gaps, kind='mergesort')

            count = 0
            while count < remainder:
                expect[order_to_adjust[count]] += piece
                count += 1

            # run chi squared test and add p values
            test_out = stats.chisquare(obs, expect)
            out_dict['p_value_from_%s' % l_form].extend(
                [test_out.pvalue for i in range(5)]
            )

        # update list index for splicing
        start_i += 5

    out_df = pd.DataFrame.from_dict(out_dict)
    out_name = out_dir + '\\landform_transitions_chi_square.csv'
    out_df.to_csv(out_name)

    return out_name


def violin_ttest(
    detrended_dem: str,
    analysis_dir: str,
    zs: Union[str, List[Union[float, int]]],
    thresh: Union[float, int] = 0.50
) -> str:
    """Creates an output csv with results from the violin t-test.

    For more information on this test, and the meaning behind it please ref:
        https://gcs-gui-documentation.readthedocs.io/en/latest/Pages/tab8.html

    Args:
        :param aligned_csv: Aligned GCS csv
        :param analysis_dir: the directory path to save /nested_analysis/OUTPUTS. 
        :param z_labels: key Z column headers in aligned_gcs.csv. 
        :param thresh: Absolute value Ws/Zs threshold to analyze (default = 0.5).

    :Returns: The path to the output .csv file. 
    """

    # prepare output dictionary
    out_dict = {
        'from stage': [],
        'from elevation': [],
        'to variable': [],
        'mean': [],
        'std': [],
        'median': [],
        'max': [],
        'min': [],
        'range': [],
        'welch_ttest_t': [],
        'welch_ttest_p': [],
    }

    # pull in aligned GCS data
    dem_dir = os.path.dirname(detrended_dem)
    gcs_dir = dem_dir + '\\gcs_tables'
    aligned_csv = gcs_dir + '\\aligned_gcs.csv'
    aligned_df = pd.read_csv(aligned_csv)

    # prepare output directory
    out_dir = analysis_dir + '\\nesting_analysis'
    if not os.path.exists(out_dir):
        os.mkdir(out_dir)

    # get units for labeling
    u = file_functions.get_label_units(detrended_dem)[0]

    # prep flow stage labels
    z_labels = [file_functions.float_keyz_format(z) + u for z in zs]

    topos = [
        'High Zs, > %s' % thresh,
        'Low Zs, < -%s' % thresh,
    ]

    for i, lower in enumerate(z_labels[:-1]):
        higher = z_labels[i + 1]
        subs = [
            aligned_df.loc[lambda aligned_df:
                           aligned_df[f'{lower}_Zs'] > thresh],
            aligned_df.loc[lambda aligned_df:
                           aligned_df[f'{lower}_Zs'] < -thresh],
        ]
        higher_col = f'{higher}_Ws'
        higher_variable = f'{higher} Ws'

        t_ins = []
        for j, topo in enumerate(topos):
            sub = subs[j]
            out_dict['from stage'].append(lower)
            out_dict['from elevation'].append(topo)
            out_dict['to variable'].append(higher_variable)
            out_dict['mean'].append(sub[higher_col].mean())
            out_dict['std'].append(sub[higher_col].std())
            out_dict['median'].append(sub[higher_col].median())
            out_dict['max'].append(sub[higher_col].max())
            out_dict['min'].append(sub[higher_col].min())
            out_dict['range'].append(
                sub[higher_col].max() - sub[higher_col].min()
            )
            t_ins.append(sub[higher_col].to_numpy())

        # run the t-test
        t, p = stats.ttest_ind(
            t_ins[0],
            t_ins[1],
            equal_var=False,
            nan_policy='omit',
        )

        out_dict['welch_ttest_t'].extend([t, t])
        out_dict['welch_ttest_p'].extend([p, p])

    out_df = pd.DataFrame.from_dict(out_dict)

    logging.info(out_df)

    thresh_label = file_functions.float_keyz_format(thresh) + u
    out_csv = out_dir + '\\preferential_nesting_ttest.csv'
    out_df.to_csv(out_csv)

    return out_csv
